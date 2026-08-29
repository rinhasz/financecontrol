"""Posição de investimentos — fase 1 do doc 16.

Objetivo da fase: **capturar a posição, não calcular nada**. O app passa a
espelhar o que o internet banking mostra, para servir de conferência e, nas
fases seguintes, orientar resgate e simular realocação.

Três origens, formatos diferentes, mesma tabela:

| origem | onde | fungível? | regime do saldo |
|---|---|---|---|
| `emissao_itau` | LCI, LCA, LIG, CDB Itaú | não | accrual |
| `acoes` | ações e ETFs | sim | mercado (MTM) |
| `rf_corretora` | CDB de outros bancos, CRI, CRA, DEB | sim | mercado (MTM) |

Produto de emissão não tem preço unitário no extrato — é o mesmo papel do
começo ao fim, e o banco informa só o saldo. Por isso `pu = 1` e
`quantidade = valor de aplicação`, como o doc 16 define: mantém a coluna
preenchida sem inventar cotação.
"""
import io
import re
import unicodedata
from datetime import datetime

from flask import Blueprint, jsonify, request

from .db import db, get_db

bp = Blueprint('investimentos', __name__)

ORIGENS = ('emissao_itau', 'acoes', 'rf_corretora')

# Rótulos das abas do arquivo de carga (doc 16)
ABAS = {
    'POSICAO_DETALHADA_EMISSAO_ITAU': 'emissao_itau',
    'POSICAO_DETALHADA_ACOES': 'acoes',
    'POSICAO_DETALHADA_RF_CORRETORA': 'rf_corretora',
}

PRODUTOS_CORRETORA = ('CDB', 'CRA', 'CRI', 'DEB', 'LCI', 'LCA', 'LIG')


def _txt(v) -> str:
    return re.sub(r'\s+', ' ', str(v if v is not None else '')).strip()


def _sem_acento(v: str) -> str:
    return unicodedata.normalize('NFKD', v).encode('ascii', 'ignore').decode().upper()


def _num(v):
    """Aceita float puro (xlsx) e texto no formato BR (HTML do internet banking)."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _txt(v).replace('R$', '').strip()
    if re.search(r'\d,\d{1,2}$', s):          # 1.234,56
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None


def _data(v):
    """dd/mm/aa, dd/mm/aaaa ou datetime. Devolve None se não der para ler —
    o arquivo de carga tem datas truncadas em algumas linhas e travar por causa
    disso perderia a posição inteira."""
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = _txt(v)
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})$', s)
    if not m:
        return None
    d, mes, a = m.groups()
    ano = int(a) if len(a) == 4 else 2000 + int(a)
    try:
        return f'{ano:04d}-{int(mes):02d}-{int(d):02d}'
    except ValueError:
        return None


def parse_remuneracao(texto: str) -> tuple:
    """`(indexador, perc_indexador, taxa)` a partir do texto do banco.

    Três formatos convivem na mesma coluna:
        '94.0000% do DI'   -> ('DI',   94.0, None)
        '13.240% aa'       -> ('PRE',  None, 13.24)
        'IPCA + 4.05% aa'  -> ('IPCA', None, 4.05)

    Separar percentual-do-índice de taxa-ao-ano importa: são coisas diferentes,
    e somar as duas num campo só impediria comparar produtos na hora de decidir
    o resgate.
    """
    s = _txt(texto)
    if not s or s == '---':
        return None, None, None

    m = re.match(r'^([\d.,]+)\s*%\s*do\s+(\w+)', s, re.I)
    if m:
        return _sem_acento(m.group(2)), _num(m.group(1)), None

    m = re.match(r'^([A-Za-z]+)\s*\+\s*([\d.,]+)\s*%', s)
    if m:
        return _sem_acento(m.group(1)), None, _num(m.group(2))

    m = re.match(r'^([\d.,]+)\s*%', s)
    if m:
        return 'PRE', None, _num(m.group(1))

    return None, None, None


def parse_ativo_corretora(rotulo: str) -> dict:
    """Extrai o que o rótulo do papel carrega.

    'CDB1223I9RK - CDB PINEBM IPCA 7% 18/01/2027' traz código, produto,
    indexador, taxa e vencimento num campo só. O arquivo de carga vem com o
    rótulo truncado, então cada pedaço é opcional: o que der para ler, lê.
    """
    s = _txt(rotulo)
    out = {'ativo': s, 'produto': None, 'indexador': None, 'taxa': None,
           'data_vencimento': None}

    codigo, _, resto = s.partition(' - ')
    if resto:
        out['ativo'] = codigo.strip()
        s = resto.strip()

    up = _sem_acento(s)
    for p in PRODUTOS_CORRETORA:
        if up.startswith(p):
            out['produto'] = p
            break

    m = re.search(r'\b(IPCA|IGPM|CDI|DI|SELIC|PRE)\b', up)
    if m:
        out['indexador'] = 'DI' if m.group(1) == 'CDI' else m.group(1)

    m = re.search(r'([\d.,]+)\s*%', s)
    if m:
        out['taxa'] = _num(m.group(1))

    m = re.search(r'(\d{2}/\d{2}/\d{4})', s)
    if m:
        out['data_vencimento'] = _data(m.group(1))

    return out


# ---------------------------------------------------------------- leitura ---

def _linhas_xlsx(conteudo: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    return {nome: list(wb[nome].iter_rows(values_only=True)) for nome in wb.sheetnames}


def _linhas_html(conteudo: bytes) -> list:
    """O internet banking exporta `.xls` que na verdade é HTML com uma <table>.

    Sem isto, os arquivos que o usuário realmente baixa não abrem — o `.xlsx` de
    carga é a exceção, não a regra.
    """
    try:
        txt = conteudo.decode('utf-8')
    except UnicodeDecodeError:
        txt = conteudo.decode('latin-1')
    linhas = []
    for tr in re.findall(r'<tr[^>]*>(.*?)</tr>', txt, re.S | re.I):
        celulas = [re.sub(r'<[^>]+>', '', c) for c in
                   re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', tr, re.S | re.I)]
        celulas = [_txt(c.replace('&nbsp;', ' ')) for c in celulas]
        if any(celulas):
            linhas.append(celulas)
    return linhas


def _classificar_html(linhas: list) -> str:
    """Ações e RF de corretora saem em arquivos diferentes com cabeçalho quase
    igual; o que separa é a coluna 'Variação do dia'."""
    cabecalho = _sem_acento(' '.join(linhas[0])) if linhas else ''
    if 'VARIACAO' in cabecalho:
        return 'acoes'
    if 'EMPRESA' in cabecalho or 'EMISSOR' in cabecalho:
        return 'rf_corretora'
    return None


# ------------------------------------------------------------------ parse ---

def parse_emissao_itau(linhas: list, data_posicao: str) -> list:
    itens = []
    for r in linhas[1:]:
        produto_bruto = _txt(r[0] if len(r) > 0 else '')
        if not produto_bruto:
            # o arquivo de carga traz um bloco final só com saldos, sem produto
            continue
        # 'LIG-DI DI' / 'CDB COFRINHOS' -> LIG / CDB
        tipo = _sem_acento(re.split(r'[-\s]', produto_bruto)[0])
        indexador, perc, taxa = parse_remuneracao(r[5] if len(r) > 5 else '')
        aplicacao = _num(r[4] if len(r) > 4 else None)
        itens.append({
            'data_posicao': data_posicao, 'origem': 'emissao_itau',
            'produto': tipo, 'ativo': produto_bruto, 'emissor': 'Itaú',
            'indexador': indexador, 'perc_indexador': perc, 'taxa': taxa,
            'data_aplicacao': _data(r[1] if len(r) > 1 else None),
            'data_vencimento': _data(r[2] if len(r) > 2 else None),
            'data_liquidez': _data(r[9] if len(r) > 9 else None),
            # não fungível: sem PU no extrato (doc 16)
            'pu': 1.0, 'quantidade': aplicacao, 'valor_aplicacao': aplicacao,
            'saldo_bruto_accrual': _num(r[7] if len(r) > 7 else None),
            'saldo_liquido_accrual': _num(r[8] if len(r) > 8 else None),
            'saldo_bruto_mtm': None, 'saldo_liquido_mtm': None,
        })
    return itens


def parse_acoes(linhas: list, data_posicao: str) -> list:
    itens = []
    for r in linhas[1:]:
        if len(r) < 6 or not _txt(r[1]):
            continue
        itens.append({
            'data_posicao': data_posicao, 'origem': 'acoes',
            'produto': 'ACAO', 'ativo': _txt(r[1]), 'emissor': _txt(r[0]),
            'indexador': None, 'perc_indexador': None, 'taxa': None,
            'data_aplicacao': None, 'data_vencimento': None,
            # ação liquida em D+2, mas a posição não informa: fica nulo em vez
            # de ser inventado (fase 1 não calcula nada)
            'data_liquidez': None,
            'pu': _num(r[3]), 'quantidade': _num(r[4]), 'valor_aplicacao': None,
            'saldo_bruto_accrual': None, 'saldo_liquido_accrual': None,
            'saldo_bruto_mtm': _num(r[5]), 'saldo_liquido_mtm': None,
        })
    return itens


def parse_rf_corretora(linhas: list, data_posicao: str) -> list:
    itens = []
    for r in linhas[1:]:
        if len(r) < 5 or not _txt(r[1]):
            continue
        info = parse_ativo_corretora(r[1])
        itens.append({
            'data_posicao': data_posicao, 'origem': 'rf_corretora',
            'produto': info['produto'] or 'RF', 'ativo': info['ativo'],
            'emissor': _txt(r[0]),
            'indexador': info['indexador'], 'perc_indexador': None,
            'taxa': info['taxa'],
            'data_aplicacao': None, 'data_vencimento': info['data_vencimento'],
            'data_liquidez': None,
            'pu': _num(r[2]), 'quantidade': _num(r[3]), 'valor_aplicacao': None,
            'saldo_bruto_accrual': None, 'saldo_liquido_accrual': None,
            'saldo_bruto_mtm': _num(r[4]), 'saldo_liquido_mtm': None,
        })
    return itens


PARSERS = {'emissao_itau': parse_emissao_itau, 'acoes': parse_acoes,
           'rf_corretora': parse_rf_corretora}


def _data_da_planilha(linhas: list) -> str:
    """Data que a planilha declara em 'Saldo atualizado até'.

    **Não é confiável.** O arquivo baixado em 28/08/2026 traz os saldos de 28/08
    (a LIG prefixada soma 39.322,82, exatamente o que o site mostrava naquele
    dia) e mesmo assim carimba 25/08/2026 em todas as linhas — o Itaú não
    atualiza essa coluna. Fica como último recurso, atrás do nome do arquivo.
    """
    for r in linhas[1:]:
        if len(r) > 3:
            d = _data(r[3])
            if d:
                return d
    return None


def _data_do_nome(nome: str) -> str:
    """Data no nome do arquivo — `POSICAO_CONSOLIDADA_RF_28082026.xlsx`.

    É a fonte mais confiável que existe aqui: o internet banking carimba o nome
    com o dia do download, que é a data real da foto, enquanto a coluna interna
    fica para trás. Sem isso, importar a posição de 28/08 a gravaria como 25/08 e
    **substituiria** a posição anterior, que é a chave de substituição por
    `(data_posicao, origem)`.
    """
    from datetime import date as _d
    for dd, mm, aaaa in re.findall(r'(\d{2})(\d{2})(\d{4})', nome):
        try:
            return _d(int(aaaa), int(mm), int(dd)).isoformat()
        except ValueError:
            continue
    return None


def ler_arquivo(nome: str, conteudo: bytes, data_informada: str = None) -> tuple:
    """`(itens, data_posicao)` a partir de um arquivo de carga."""
    ext = nome.rsplit('.', 1)[-1].lower()
    # o que o usuário digitou vence tudo; depois o nome do arquivo; a coluna da
    # planilha só entra se as duas faltarem, porque ela atrasa (ver acima)
    itens, data = [], data_informada or _data_do_nome(nome)

    if ext in ('xlsx', 'xlsm'):
        abas = _linhas_xlsx(conteudo)
        for nome_aba, linhas in abas.items():
            origem = ABAS.get(_sem_acento(nome_aba).replace(' ', '_'))
            if not origem or not linhas:
                continue
            if origem == 'emissao_itau':
                data = data or _data_da_planilha(linhas)
            itens.append((origem, linhas))
    else:
        linhas = _linhas_html(conteudo)
        origem = _classificar_html(linhas)
        if origem:
            itens.append((origem, linhas))

    if not data:
        from datetime import date
        data = date.today().isoformat()

    return [(o, PARSERS[o](l, data)) for o, l in itens], data


# --------------------------------------------------------------- endpoints ---

@bp.route('/investimentos/importar', methods=['POST'])
def importar():
    arquivo = request.files.get('file')
    if not arquivo:
        return jsonify({'ok': False, 'msg': 'Nenhum arquivo enviado'}), 400

    nome = arquivo.filename or 'posicao'
    try:
        blocos, data_posicao = ler_arquivo(nome, arquivo.read(),
                                           request.form.get('data_posicao') or None)
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'Não consegui ler o arquivo: {e}'}), 400

    if not blocos:
        return jsonify({'ok': False, 'msg': 'Nenhuma posição reconhecida no arquivo'}), 400

    origens = [o for o, itens in blocos if itens]
    total = sum(len(itens) for _o, itens in blocos)
    if not total:
        return jsonify({'ok': False, 'msg': 'Arquivo reconhecido, mas sem linhas de posição'}), 400

    try:
        _gravar_posicao(blocos, data_posicao, nome, origens)
    except Exception as e:
        # erro de gravação vira mensagem na tela: um 500 sem corpo deixava o
        # usuário sem saber o que houve
        return jsonify({'ok': False, 'msg': f'Não consegui gravar a posição: {e}'}), 500

    rotulos = {'emissao_itau': 'emissão Itaú', 'acoes': 'ações',
               'rf_corretora': 'RF corretora'}
    detalhe = ', '.join(f'{len(i)} de {rotulos[o]}' for o, i in blocos if i)
    return jsonify({'ok': True, 'data_posicao': data_posicao, 'total': total,
                    'msg': f'Posição de {data_posicao}: {detalhe}'})


def _gravar_posicao(blocos, data_posicao, nome, origens):
    with db() as conn:
        cur = conn.execute(
            'INSERT INTO carga_investimento (data_posicao, arquivo, origens) VALUES (?,?,?)',
            (data_posicao, nome, ','.join(origens)))
        carga_id = cur.lastrowid

        # Substitui a posição daquela data **só nas origens presentes no
        # arquivo**: importar as ações não pode apagar a renda fixa, que veio
        # de outro arquivo.
        for origem, itens in blocos:
            if not itens:
                continue
            # A memória de valorização aponta para estas linhas; sem apagá-la
            # antes, o DELETE falha por chave estrangeira. Ela é descartável de
            # propósito: é derivada da posição, e a posição acabou de mudar —
            # "Atualizar Posições" a reconstrói do dado de mercado guardado.
            conn.execute(
                'DELETE FROM valorizacao WHERE investimento_id IN '
                '(SELECT id FROM investimento WHERE data_posicao=? AND origem=?)',
                (data_posicao, origem))
            conn.execute('DELETE FROM investimento WHERE data_posicao=? AND origem=?',
                         (data_posicao, origem))
            conn.executemany(
                'INSERT INTO investimento (data_posicao, origem, produto, ativo, emissor, '
                'indexador, perc_indexador, taxa, data_aplicacao, data_vencimento, '
                'data_liquidez, pu, quantidade, valor_aplicacao, saldo_bruto_accrual, '
                'saldo_liquido_accrual, saldo_bruto_mtm, saldo_liquido_mtm, carga_id) '
                'VALUES (:data_posicao,:origem,:produto,:ativo,:emissor,:indexador,'
                ':perc_indexador,:taxa,:data_aplicacao,:data_vencimento,:data_liquidez,'
                ':pu,:quantidade,:valor_aplicacao,:saldo_bruto_accrual,'
                ':saldo_liquido_accrual,:saldo_bruto_mtm,:saldo_liquido_mtm,:carga_id)',
                [{**i, 'carga_id': carga_id} for i in itens])


@bp.route('/investimentos/atualizar', methods=['POST'])
def atualizar_posicoes():
    """Valoriza a última posição até hoje, dia útil a dia útil (doc 16 §fase 2)."""
    from . import valorizacao
    res = valorizacao.atualizar()
    return jsonify(res), (200 if res.get('ok') else 400)


@bp.route('/investimentos/<int:iid>/memoria')
def memoria_calculo(iid):
    """Passo a passo de como o saldo daquele papel chegou onde chegou."""
    from . import valorizacao
    return jsonify(valorizacao.memoria(iid))


@bp.route('/investimentos/datas')
def datas():
    conn = get_db()
    rows = conn.execute(
        'SELECT data_posicao, COUNT(*) n, SUM(COALESCE(saldo_bruto_mtm, saldo_bruto_accrual)) total '
        'FROM investimento GROUP BY data_posicao ORDER BY data_posicao DESC').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@bp.route('/investimentos')
def listar():
    """Posição detalhada e consolidada de uma data.

    O `saldo` de cada linha é `MTM` quando existe, `accrual` quando não —
    é o valor da posição hoje, que é o que o internet banking mostra e o que
    esta tela existe para conferir.
    """
    conn = get_db()
    data = request.args.get('data')
    if not data:
        row = conn.execute('SELECT MAX(data_posicao) d FROM investimento').fetchone()
        data = row['d']
    if not data:
        conn.close()
        return jsonify({'data_posicao': None, 'itens': [], 'consolidado': {}, 'total': 0})

    # `saldo` é o valorizado quando existe: é a posição de hoje, que é o que a
    # tela deve mostrar. `saldo_posicao` fica ao lado para dar para comparar com
    # o que o banco informou na data da foto.
    rows = conn.execute(
        'SELECT *, COALESCE(saldo_bruto_mtm, saldo_bruto_accrual) as saldo_posicao, '
        'COALESCE(saldo_valorizado, saldo_bruto_mtm, saldo_bruto_accrual) as saldo '
        'FROM investimento WHERE data_posicao=? ORDER BY origem, produto, ativo',
        (data,)).fetchall()
    itens = [dict(r) for r in rows]
    conn.close()

    def agrupar(chave):
        """Soma cru e arredonda uma vez só. Arredondar a cada parcela fazia o
        total do grupo divergir da soma das suas linhas por frações de centavo —
        invisível em reais, mas suficiente para a quebra por indexador não
        fechar com o grupo que a contém."""
        out = {}
        for i in itens:
            k = i.get(chave) or 'Não informado'
            g = out.setdefault(k, {'chave': k, 'total': 0.0, 'itens': 0})
            g['total'] += i['saldo'] or 0
            g['itens'] += 1
        for g in out.values():
            g['total'] = round(g['total'], 2)
        return sorted(out.values(), key=lambda g: -g['total'])

    datas_val = {i['data_valorizacao'] for i in itens if i['data_valorizacao']}
    return jsonify({
        'data_posicao': data,
        'data_valorizacao': max(datas_val) if datas_val else None,
        'itens': itens,
        'total': round(sum(i['saldo'] or 0 for i in itens), 2),
        'total_posicao': round(sum(i['saldo_posicao'] or 0 for i in itens), 2),
        'consolidado': {
            'produto': agrupar('produto'),
            'indexador': agrupar('indexador'),
            'emissor': agrupar('emissor'),
            'origem': agrupar('origem'),
        },
    })
